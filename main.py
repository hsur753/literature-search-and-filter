import sys
import re
import hashlib
import argparse
import bibtexparser
from openpyxl import Workbook
from openpyxl import load_workbook
from langdetect import detect, DetectorFactory

databases = {"acm": "ACM Digital Library", "scdrt": "ScienceDirect", "ieee": "IEEE Explore"}

parser = argparse.ArgumentParser()
parser.add_argument("--file_path", required=True,  help = "Path to Bibtex File")
parser.add_argument("--mode", required=True,  help = str(databases))
parser.add_argument("--page_limit", required=True,  help = "Minimum Page Count")
parser.add_argument("--excel_path", help = "Path to Excel File")
parser.add_argument('--search_keywords', action='store_true', help = "Keywords to search in Title, Abstract")
args = parser.parse_args()

PATH_TO_FILE = args.file_path
MODE = args.mode
PAGE_LIMIT = (int) (args.page_limit)
EXCEL_PATH = args.excel_path
SEARCH_KEYWORDS = args.search_keywords

#PROCESS AND FILTER BIBTEX ARTICLES
def process_bibtex(bibtex_data):

    #TRANSFORM PAGE RANGE TO PAGE COUNT
    condition = lambda entry: "pages" in entry and "-" in entry["pages"]
    processed_bibtex_list = [entry for entry in bibtex_data if condition(entry)]
    for entry in processed_bibtex_list:
        entry["pages"] = get_page_count(entry["pages"])

    return processed_bibtex_list

def apply_filter(bibtex_list, page_key, title_key):

    #ENFORCE CONSISTENT LANGUAGE DETECTION
    DetectorFactory.seed = 0

    #FILTER CRITERIA
    condition = lambda entry: page_key in entry and (int) (entry[page_key]) >= PAGE_LIMIT and detect(entry[title_key]) == "en"
    full_paper_list = [entry for entry in bibtex_list if condition(entry)]

    #REMOVE DUPLICATES
    for paper in full_paper_list:
        paper["ID"] = hashlib.md5(paper[title_key].encode()).hexdigest()
    unique_papers_list = list({entry["ID"]: entry for entry in full_paper_list}.values())

    return unique_papers_list

def get_page_count(value):
    index = value.find("-")
    try:
        return (int) (value[index + 1:]) - (int) (value[0 : index])
    except:
        return -1

#EXCEL WORKBOOK FUNCTIONS
def initialise_workbook(references_list, mode):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = databases[mode]
    write_to_sheet(sheet, references_list)
    return workbook

def edit_workbook(source, references_list, mode):  
    workbook = load_workbook(source)
    sheet = workbook.create_sheet(databases[mode])
    write_to_sheet(sheet, references_list)
    return workbook

def write_to_sheet(sheet, references_list):
    sheet.append(["ID", "Author", "Title", "DOI"])
    for row in references_list:
        sheet.append([row["ID"], row.get("author", "N/A"), row["title"], row.get("url", row.get("doi", "N/A"))])

#APPLY FILTER BASED ON DB SOURCE
def run_filter(mode):
    print(len(bib_database.entries))
    match mode:
        case "acm":
            return apply_filter(bib_database.entries, "numpages", "title")
        case "scdrt" | "ieee":
            processed_bibtex_list = process_bibtex(bib_database.entries)
            return apply_filter(processed_bibtex_list, "pages", "title")

def search_keywords(keywords, text):

    is_exact_match = lambda word: word[0] == "\"" and word[-1] == "\""
    run_exact_search = lambda word: re.search(f"\\b{word[1:-1]}(\\b|\\W)", text, re.IGNORECASE) is not None
    run_estimate_search = lambda word: re.search(f"\\b{word}\\w*(\\b|\\W)", text, re.IGNORECASE) is not None

    keyword_logic_list = []
    for keyword_set in keywords:
        or_operature_list = [run_exact_search(keyword) if is_exact_match(keyword) else run_estimate_search(keyword) for keyword in keyword_set]
        keyword_logic_list.append(any(or_operature_list))

    return all(keyword_logic_list)

def run_search(references_list, raw_keywords):
    keywords_list = [[word.strip() for word in line.split(",")] for line in raw_keywords]
    extract_text = lambda entry: " ".join([entry.get("abstract", ""), entry.get("title", ""), entry.get("keywords", "")])
    filtered_list = [entry for entry in references_list if search_keywords(keywords_list, extract_text(entry))]
    return filtered_list


if __name__ == '__main__':

    print("Processing...")
    with open(PATH_TO_FILE, encoding='utf8') as bibtex_file:
        bib_database = bibtexparser.load(bibtex_file)

    references_list = run_filter(MODE)

    if (SEARCH_KEYWORDS):
        print("\nEnter required keywords seperated by a comma (,) to run OR operator.\n" +
        "Add keywords on a newline to run AND operator or type \"X\" to finish.\n" +
        "To search for an exact match of a keyword enclose it within quotes (\"\")\n")

        raw_keywords = []
        for line in sys.stdin:
            if 'X' == line.strip():
                break
            else:
                raw_keywords.append(line.strip())
        
        references_list = run_search(references_list, raw_keywords)

    if (EXCEL_PATH is None):
        EXCEL_PATH = "filtered_papers.xlsx" 
        reference_library = initialise_workbook(references_list, MODE)
    else:
        reference_library = edit_workbook(EXCEL_PATH, references_list, MODE)

    reference_library.save(filename=EXCEL_PATH)

    print()
    print("--------------------------------------------------------------------")
    print(f"{len(bib_database.entries)} Papers Found")
    print(f"{len(references_list)} Papers Saved After Filtering")
    print("--------------------------------------------------------------------")